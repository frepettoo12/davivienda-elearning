"""
Seed de la colección `companies` (multi-tenant).

Crea/actualiza companies/davivienda con los valores que hasta ahora estaban
hardcodeados en el código (branding, dominios, prompts, defaults de voz/avatar).
También migra el shell SCORM global (config/scorm.shell_html) al doc de la company.

Idempotente: usa set(merge=True). Correr local con ADC:
    python3 scripts/seed_companies.py [--project davivienda-elearning] [--dry-run]

FUENTE EXCEL (recomendado): las empresas se administran en scripts/companies.xlsx
(una fila por empresa; listas separadas por coma). Editar el Excel y correr:
    python3 scripts/seed_companies.py --excel scripts/companies.xlsx
Generar/regenerar el template con la fila de Davivienda:
    python3 scripts/seed_companies.py --write-template

Alternativas: bloque COMPANIES embebido, o JSON con --file.
"""
import argparse
import json
import sys

import firebase_admin
from firebase_admin import credentials, firestore

EXCEL_COLUMNS = [
    ("company_id", "ID (slug, ej: davivienda)"),
    ("nombre", "Nombre"),
    ("activo", "Activo (si/no)"),
    ("dominios", "Dominios de email (coma)"),
    ("learning_domains", "Dominios equipo Learning (coma)"),
    ("industria", "Industria"),
    ("descripcion_prompt", "Descripción para prompts IA"),
    ("color_primario", "Color primario (#hex)"),
    ("color_acento", "Color acento (#hex)"),
    ("logo_url", "Logo URL (https)"),
    ("fuente_titulos", "Fuente títulos"),
    ("fuente_texto", "Fuente texto"),
    ("email_from_name", "Nombre remitente emails"),
    ("app_url", "URL del frontend"),
    ("areas", "Áreas (coma)"),
    ("lms_nombre", "LMS destino"),
    ("voice_id", "Voz ElevenLabs (ID)"),
    ("avatar_id", "Avatar HeyGen (ID)"),
    ("passing_score", "Puntaje aprobación (%)"),
]


def _split(v):
    return [s.strip() for s in str(v).split(",") if s and str(s).strip()] if v else []


def _row_to_company(row: dict) -> tuple[str, dict]:
    """Convierte una fila del Excel al doc de companies/{id}."""
    import re
    cid = str(row.get("company_id") or "").strip().lower()
    if not re.fullmatch(r"[a-z0-9][a-z0-9_-]*", cid or ""):
        # Vacío o no-slug (ej: la fila de ayuda del template) → se saltea.
        raise ValueError(f"company_id inválido: {cid!r}")
    nombre = str(row.get("nombre") or cid.title()).strip()
    dominios = _split(row.get("dominios"))
    return cid, {
        "nombre": nombre,
        "activo": str(row.get("activo") or "si").strip().lower() not in ("no", "false", "0"),
        "dominios": dominios,
        "learning_domains": _split(row.get("learning_domains")) or dominios,
        "industria": str(row.get("industria") or "").strip(),
        "descripcion_prompt": str(row.get("descripcion_prompt") or "").strip(),
        "branding": {
            "nombre_display": f"{nombre} E-Learning",
            "color_primario": str(row.get("color_primario") or "#DA291C").strip(),
            "color_acento": str(row.get("color_acento") or "#FFD700").strip(),
            "logo_url": (str(row.get("logo_url")).strip() or None) if row.get("logo_url") else None,
            "fuente_titulos": str(row.get("fuente_titulos") or "Montserrat").strip(),
            "fuente_texto": str(row.get("fuente_texto") or "Open Sans").strip(),
        },
        "email": {"from_name": str(row.get("email_from_name") or f"{nombre} E-Learning").strip()},
        "app_url": (str(row.get("app_url")).strip() or None) if row.get("app_url") else None,
        "areas": _split(row.get("areas")) or None,
        "lms_nombre": str(row.get("lms_nombre") or "").strip() or None,
        "defaults": {
            "voice_id": str(row.get("voice_id") or "JddqVF50ZSIR7SRbJE6u").strip(),
            "avatar_id": str(row.get("avatar_id") or "Hada_LivelyGestures_Front_public").strip(),
            "passing_score": int(row.get("passing_score") or 70),
        },
        "scorm": {"shell_html": "", "manifest_identifier": f"{cid}_scorm"},
    }


def load_excel(path: str) -> dict:
    """Lee companies.xlsx (fila 1 = headers con las keys de EXCEL_COLUMNS)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    companies = {}
    for xrow in ws.iter_rows(min_row=2, values_only=True):
        row = {h: v for h, v in zip(headers, xrow) if h}
        if not any(v not in (None, "") for v in row.values()):
            continue
        try:
            cid, data = _row_to_company(row)
        except ValueError:
            continue  # fila de ayuda / incompleta
        companies[cid] = data
    return companies


def write_template(path: str) -> None:
    """Genera companies.xlsx con headers + la fila de Davivienda."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    for i, (key, label) in enumerate(EXCEL_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=key)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        ws.cell(row=2, column=i, value=label).font = Font(italic=True, size=9, color="888888")
        ws.column_dimensions[c.column_letter].width = max(18, len(label) + 2)
    # Fila 2 = ayuda (se ignora al importar porque no tiene company_id)
    davi = COMPANIES["davivienda"]
    fila = {
        "company_id": "davivienda",
        "nombre": davi["nombre"],
        "activo": "si",
        "dominios": ", ".join(davi["dominios"]),
        "learning_domains": ", ".join(davi["learning_domains"]),
        "industria": davi["industria"],
        "descripcion_prompt": davi["descripcion_prompt"],
        "color_primario": davi["branding"]["color_primario"],
        "color_acento": davi["branding"]["color_acento"],
        "logo_url": davi["branding"]["logo_url"] or "",
        "fuente_titulos": davi["branding"]["fuente_titulos"],
        "fuente_texto": davi["branding"]["fuente_texto"],
        "email_from_name": davi["email"]["from_name"],
        "app_url": davi["app_url"] or "",
        "areas": ", ".join(davi["areas"]),
        "lms_nombre": davi["lms_nombre"],
        "voice_id": davi["defaults"]["voice_id"],
        "avatar_id": davi["defaults"]["avatar_id"],
        "passing_score": davi["defaults"]["passing_score"],
    }
    for i, (key, _label) in enumerate(EXCEL_COLUMNS, start=1):
        ws.cell(row=3, column=i, value=fila.get(key, ""))
    wb.save(path)
    print(f"✓ Template escrito en {path} (fila 3 = Davivienda; agregar una fila por empresa)")

COMPANIES = {
    "davivienda": {
        "nombre": "Davivienda",
        "activo": True,
        # Cualquier dominio acá = miembro de la empresa (resuelve el tenant al loguear).
        "dominios": ["davivienda.com", "alkemy.org"],
        # Subset con rol "learning" (staff del dashboard). El resto de dominios
        # de la empresa (si los hubiera) serían solicitantes.
        "learning_domains": ["davivienda.com", "alkemy.org"],
        # Contexto para prompts GPT (reemplaza "Davivienda (banco colombiano)").
        "industria": "banca colombiana",
        "descripcion_prompt": "banco colombiano líder en servicios financieros",
        "branding": {
            "nombre_display": "Davivienda E-Learning",
            "color_primario": "#DA291C",
            "color_acento": "#FFD700",
            "logo_url": None,
            "fuente_titulos": "Montserrat",
            "fuente_texto": "Open Sans",
        },
        "email": {
            "from_name": "Davivienda E-Learning",
        },
        # Base del frontend para links en emails (None = usa APP_URL global).
        "app_url": None,
        "areas": [
            "Banca Personal", "Banca Empresarial", "Banca Corporativa",
            "Operaciones", "Tecnología", "Riesgos", "Cumplimiento",
            "Talento Humano", "Mercadeo", "Jurídica", "Auditoría", "Otra",
        ],
        "lms_nombre": "Territorium",
        "defaults": {
            "voice_id": "JddqVF50ZSIR7SRbJE6u",  # valeria
            "avatar_id": "Hada_LivelyGestures_Front_public",  # hada
            "passing_score": 70,
        },
        "scorm": {
            "shell_html": "",  # se migra desde config/scorm si existe
            "manifest_identifier": "davivienda_scorm",
        },
    },
}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", default="davivienda-elearning")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--file", help="JSON extra con {company_id: {...}} para sumar/pisar")
    parser.add_argument("--excel", help="Excel de empresas (scripts/companies.xlsx) como fuente")
    parser.add_argument("--write-template", action="store_true",
                        help="Genera scripts/companies.xlsx con headers + fila Davivienda y sale")
    args = parser.parse_args()

    if args.write_template:
        write_template(args.excel or "scripts/companies.xlsx")
        return

    if args.excel:
        companies = load_excel(args.excel)
        if not companies:
            print("El Excel no tiene filas de empresa válidas — nada que hacer.")
            return
    else:
        companies = dict(COMPANIES)
    if args.file:
        with open(args.file) as f:
            companies.update(json.load(f))

    if not firebase_admin._apps:
        firebase_admin.initialize_app(
            credentials.ApplicationDefault(), {"projectId": args.project}
        )
    db = firestore.client()

    # Migrar shell SCORM global existente al doc de davivienda (solo si no lo tiene).
    legacy_shell = ""
    cfg = db.collection("config").document("scorm").get()
    if cfg.exists:
        legacy_shell = (cfg.to_dict() or {}).get("shell_html", "") or ""
    if legacy_shell and "davivienda" in companies and not companies["davivienda"]["scorm"].get("shell_html"):
        existing = db.collection("companies").document("davivienda").get()
        already = ""
        if existing.exists:
            already = ((existing.to_dict() or {}).get("scorm") or {}).get("shell_html", "") or ""
        if not already:
            companies["davivienda"]["scorm"]["shell_html"] = legacy_shell
            print(f"→ Migrando config/scorm.shell_html ({len(legacy_shell)} chars) a companies/davivienda")

    for cid, data in companies.items():
        # No pisar un shell ya guardado (via endpoint scorm_shell) con vacío al re-correr.
        if not data.get("scorm", {}).get("shell_html"):
            data.get("scorm", {}).pop("shell_html", None)
        if args.dry_run:
            shell_len = len(data.get("scorm", {}).get("shell_html", ""))
            print(f"[dry-run] companies/{cid}:")
            print(json.dumps({**data, "scorm": {**data.get("scorm", {}), "shell_html": f"<{shell_len} chars>"}}, indent=2, ensure_ascii=False))
            continue
        data["updated_at"] = firestore.SERVER_TIMESTAMP
        db.collection("companies").document(cid).set(data, merge=True)
        print(f"✓ companies/{cid} seeded ({data['nombre']})")

    if args.dry_run:
        print("(dry-run: no se escribió nada)")


if __name__ == "__main__":
    sys.exit(main())
